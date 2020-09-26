import xlrd2,win32ui,xlwt
import openpyxl
import xlutils
import os.path
import copy

inputFile = '结果'
outputPath = ''
outputFile =''
workpath =str(os.getcwd()) + "\Table\\"


def __int__():
    print("O1K")
__int__()
#选择文件
def SelectFile(path = ''):
    dlg = win32ui.CreateFileDialog(1)  # 1表示打开文件对话框
    dlg.SetOFNInitialDir(path)  # 设置打开文件对话框中的初始显示目录
    flag =  dlg.DoModal()
    if flag == 1:
        filename = dlg.GetPathName()  # 获取选择的文件名称
        return filename
    else:
        print("取消")
        return 0

import tkinter.filedialog as tkDlg
# def OpenFile(path=''):
#     dlg = tkDlg.file
    

# 通用选择文件
def SaveXlsxFile (path=''):
    dlg = tkDlg.asksaveasfilename(title = '选择输出文件', filetypes=[("xlsx",".xlsx")])
    # dlg =  dlg.replace("/","\\")
    print(dlg)
    return dlg

def Replace(value,oldChar,newChar):
    value

# SelecFileCom()
#读取并输出excel文件为二维 list 
def ReadExel(path,name='',sheet_name=''):
    full_path = ''
    if name == '':
         #定义输出路径
        full_path = path
    else:
        full_path = path  + name

    print("准备读取："+full_path)
    res = []
    if(os.path.isfile(full_path)):
        #读取excel文件
        data = xlrd2.open_workbook(full_path)
        #读取所有sheet名称
        sheet = data.sheet_names()
        
        if sheet_name !='':
             
            table = data.sheet_by_name(sheet_name)
        else:
            # 读取第一个sheet表格
            table = data[0]
        
        if len(sheet)>0:
            for row in range(table.nrows) :
                res.append(table.row_values(row))
                # print(table.row_values(row))
            print( full_path + "读取完毕")
            return res
        else:
            print('文件中无数据')
            return 1
    else:
        print("文件不存在,导入失败")
        return 2

#根据新列顺序，转换list
def Exdata(old_list,newcol):
    new_list = copy.deepcopy(old_list)
    print(new_list[0][0])
    index = len(old_list)

    i=0
    for row in old_list:
        # print(row)
        for a in range(len(row)):
            # print(row)
            # print(row[int(newcol[a])])
            print(str(a)+","+row[int(newcol[a])])
            print(new_list[0][0])
            new_list[i][a] = str(row[int(newcol[a])]) 
            # print(new_list[i][a])
        # print(new_list[i])
        i+=1
    # print(new_list)
    return new_list


def WriteExcel_xls_row(path,name,sheet_name,table): 
    
    index = len(table)
    fullpath = path + name

    workbook = xlwt.Workbook()

    # 创建一个sheet
    newTable = workbook.add_sheet(sheet_name)

    for r in range(0, index):
        # print("r"+ str(r))
        for c in range(0, len(table[r])):
            # print(str(r)+'-'+ str(c))
            newTable.write(r , c , table[r][c])
    if os.path.isdir(path):
        workbook.save(fullpath)
        print("xls格式表格写入数据成功！")
        print('输出路径为：'+ fullpath)
    else:
        print("路径不存在:"+ path)



#通过openpyxl 输出xlsx文件
def WriteExcel_xlsx_row(path,table,name='',sheet_name=''):
    #获取 行数
    index = len(table)
    fullpath = ''
    if name == '':
         #定义输出路径
        fullpath = path + '.xlsx'
    else:
        fullpath = path  + name
    print('待输出文件：'+fullpath)
    if fullpath!='':
        print('开始输出excel文件,输出路径为：'+ fullpath)
        # 创建一个文件和sheet
        workbook = openpyxl.Workbook()

        workbook.active
        # print('s')
        workbook.remove(workbook['Sheet'])
        newTable = workbook.create_sheet(sheet_name)
        # print(newTable)
        # newTable = workbook[sheet_name]

        for r in range(0, index):
            # print("r"+ str(r))
            for c in range(0, len(table[r])):
                print(str(r)+'-'+ str(c))
                newTable.cell(r+1,c+1,table[r][c])
        workbook.save(fullpath)
        print("xls格式表格写入数据成功！")
    else:
        print("路径不存在:"+ fullpath)

# #输出例子
# value3 = [["姓名", "性别", "年龄", "城市", "职业"],
#           ["111", "女", "66", "石家庄", "运维工程师"],
#           ["222", "男", "55", "南京", "饭店老板"],
#           ["333", "女", "27", "苏州", "保安1"],]
# value2= Exdata(value3, [2,3,4,0,1])
# WriteExcel_xls_rowx(workpath,'3.xlsx','Sheet1',value2)
# WriteExcel_xls_row(workpath,'1.xls','sheet',value3)


# #输出2.xlsx 文件 
# WriteExcel_xls_rowx(workpath,'2.xlsx','Sheet1',value3)


# #例子2
# a = ReadExel(workpath,'test.xlsx','Sheet1')
# WriteExcel_xls_rowx(workpath,'2.xlsx','Sheet1',a)