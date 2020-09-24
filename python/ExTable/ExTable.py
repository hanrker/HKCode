import xlrd2,win32ui,xlwt
import openpyxl
import xlutils
import os.path
import copy

inputFile = '结果'
outputPath = ''
outputFile =''
workpath =str(os.getcwd()) + "\ExTable\\"


#选择文件
def SelectFile():
    dlg = win32ui.CreateFileDialog(1)  # 1表示打开文件对话框
    dlg.SetOFNInitialDir(workpath)  # 设置打开文件对话框中的初始显示目录
    dlg.DoModal()
    filename = dlg.GetPathName()  # 获取选择的文件名称
    print(filename)
    return filename

inputFile = SelectFile()
# print(inputFile)

#读取并输出excel文件为二维 list 
def ReadExel(path,name,sheet_name):
    full_path = path + name
    print(full_path)
    res = []
    if(os.path.isfile(full_path)):
        #读取excel文件
        data = xlrd2.open_workbook(full_path)
        #读取所有sheet名称
        sheet = data.sheet_names()
        
         # 读取第一个sheet表格
        table = data.sheet_by_name(sheet_name)

        if len(sheet)>0:
            for row in range(table.nrows) :
                res.append(table.row_values(row))
                # print(table.row_values(row))
            print("读取完毕")
        else:
             print('Please select file')
    else:
        print("导入失败")
    
    print(res[0][0])
    #输出list
    return res

a = ReadExel(workpath,'test.xlsx','Sheet1')
WriteExcel_xls_rowx(workpath,'2.xlsx','Sheet1',a)

#转换list
def Exdata(list,newcol):
    new_list = copy.deepcopy(list)
    index = len(list)

    i=0
    for row in list:
        print(row)
        for a in range(len(row)):
            print(row[newcol[a]])
            new_list[i][a] = row[newcol[a]]
            # print(new_list[i][a])
        # print(new_list[i])
        i+=1
    # print(new_list)
    return new_list


#s输出例子
value3 = [["姓名", "性别", "年龄", "城市", "职业"],
          ["111", "女", "66", "石家庄", "运维工程师"],
          ["222", "男", "55", "南京", "饭店老板"],
          ["333", "女", "27", "苏州", "保安1"],]
value2= Exdata(value3, [2,3,4,0,1])
WriteExcel_xls_rowx(workpath,'3.xlsx','Sheet1',value2)


def WriteExcel_xls_row(path,name,sheet_name,table): 
    index = len(table)
    fullpath = path + name

    workbook = xlwt.Workbook()

    # 创建一个sheet
    newTable = workbook.add_sheet(sheet_name)
    # print(table[0][1])
    # if os.path.isfile(fullpath):
    #     #复制文件
    #     workbook = xlrd2.open_workbook(fullpath)
    #     workbook.save(name)
    # else:
    #     #创建新文件
    #     workbook = xlutils.workbook(encoding = 'utf-8')
    for r in range(0, index):
        # print("r"+ str(r))
        for c in range(0, len(table[r])):
            # print(str(r)+'-'+ str(c))
            newTable.write(r , c , table[r][c])
    if os.path.isdir(path):
        workbook.save(fullpath)
        print("xls格式表格写入数据成功！")
        print('输出路径为：'+ fullpath)
    else:
        print("路径不存在:"+ path)
WriteExcel_xls_row(workpath,'1.xls','sheet',value3)


#通过openpyxl 输出xlsx文件
def WriteExcel_xls_rowx(path,name,sheet_name,table):
    #获取 行数
    index = len(table)

    #定义输出路径
    fullpath = path + name

    # 创建一个文件和sheet
    workbook = openpyxl.Workbook()

    workbook.active
    print('s')
    workbook.remove(workbook['Sheet'])
    newTable = workbook.create_sheet(sheet_name)
    print(newTable)
    # newTable = workbook[sheet_name]

    for r in range(0, index):
        # print("r"+ str(r))
        for c in range(0, len(table[r])):
            print(str(r)+'-'+ str(c))
            newTable.cell(r+1,c+1,table[r][c])
    if os.path.isdir(path):
        workbook.save(fullpath)
        print("xls格式表格写入数据成功！")
        print('输出路径为：'+ fullpath)
    else:
        print("路径不存在:"+ path)
#输出2.xlsx 文件 
WriteExcel_xls_rowx(workpath,'2.xlsx','Sheet1',value3)